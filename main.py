#Autores: Gabriela Goto, Paulo Henrique
#Data: 05/06/2022
#Quarta Avaliação

import random
from openpyxl import Workbook, load_workbook



def configuracoes():
    try:
        wb = load_workbook(filename='banco.xlsx')
        config = wb['configurações']
    except: #CASO NÃO TENHA O ARQUIVO ELE CRIA UMA CONFIGURAÇÃO BASICA
        wb = Workbook()
        config = wb.create_sheet('configurações')
        
        config.cell(column=1, row=1, value='Linhas')
        config.cell(column=2, row=1, value=3)
        config.cell(column=1, row=2, value='Colunas')
        config.cell(column=2, row=2, value=3)
        config.cell(column=1, row=3, value='Dificuldade')
        config.cell(column=2, row=3, value=2)

    dificuldade = config.cell(column=2, row=3).value
    linhas = config.cell(column=2, row=1).value
    colunas = config.cell(column=2, row=2).value
    
    wb.save('banco.xlsx')

    return int(dificuldade), int(linhas), int(colunas)



def alterarConfiguracao():
    try:
        wb = load_workbook(filename='banco.xlsx')
        config = wb['configurações']
    except:
        wb = Workbook()
        config = wb.create_sheet('configurações',1)


    while True:
        print('''Níveis de dificuldade:
        1 - Fácil
        2 - Médio
        3 - Difícil''')
        try:
            dificuldade = int(input('Qual o nível de dificuldade: '))
            linhas = int(input('Quantidade de linhas: '))
            colunas = int(input('Quantidade de colunas: '))
        except:
            print('Por favor preencha somente com numeros')
            alterarConfiguracao()
        
        if dificuldade >= 1 and dificuldade <=3 and linhas >= 3 and colunas >= 3:
            print('CONFIGURAÇÕES SALVAS COM SUCESSO\n')
            break
        else:
            print('''\033[1mConfigurações Minimas:\nDificuldade tem que ser maior de 1\nQuantidade de linhas e colunas tem que ser maior que 3\033[0m\n''')

    #utilizando esse metodo para 'forçar' os locais exatos
    config.cell(column=1, row=1, value='Linhas')
    config.cell(column=2, row=1, value=linhas)
    config.cell(column=1, row=2, value='Colunas')
    config.cell(column=2, row=2, value=colunas)
    config.cell(column=1, row=3, value='Dificuldade')
    config.cell(column=2, row=3, value=dificuldade)
    wb.save('banco.xlsx')
    
 
    menu()



def bombinhas():
    bombas = []
    dificuldade, linhas, colunas = configuracoes()
    totalCasas = (linhas*colunas)
    # NIVEL DE DIFICULDADE
    if dificuldade == 1:
        qtdBombas = int(totalCasas*0.15)  # int => para 'truncar' o numero
    elif dificuldade == 2:
        qtdBombas = int(totalCasas*0.30)
    else:
        qtdBombas = int(totalCasas*0.50)

    # ALETORIEDADE DAS BOMBAS
    while True:
        num = random.randint(1, totalCasas)
        if num not in bombas:
            bombas.append(num)
        if len(bombas) == qtdBombas: 
            break

    return bombas

def calcularAdjacente(matriz):
    dificuldade, qtdLinha, qtdColuna = configuracoes()

    for linha in range(0, qtdLinha):
        for coluna in range(0, qtdColuna):
            if matriz[linha][coluna] == '-1': 
                continue
            cont = 0
                
            for i in range(linha-1 if linha>0 else 0, linha+2 if linha<(qtdLinha-1) else qtdLinha):
                for j in range(coluna-1 if coluna>0 else 0, coluna+2 if coluna<(qtdColuna-1) else qtdColuna):
                    
                    if matriz[i][j] == '-1': 
                        cont += 1

            matriz[linha][coluna] = str(cont)
        #print(matriz)
    return matriz

def gravarTabuleiro(jogo):
    try:
        wb = load_workbook(filename='banco.xlsx', read_only=False)
    except:
        wb = Workbook()

    try:
        abaJogo = wb['jogo']
    
    except:
        abaJogo = wb.create_sheet('jogo',2)
    #resolvendo o problema de quando muda as config para um tabuleiro menor
    abaJogo.delete_rows(1, abaJogo.max_row)

     # GRAVANDO TABULEIRO
    for i in range(1, len(jogo)+1):
        for j in range(1, len(jogo[0])+1):
            abaJogo.cell(row=i, column=j, value=jogo[i-1][j-1])
   
    wb.save('banco.xlsx')

def criarTabuleiro(qtdLinha, qtdColuna):
    
    cont = 1
    bombas = bombinhas()
    game = []

    #criando primeiro uma matriz para facilitar minha vida
    for i in range(1, qtdLinha+1):
        jogo = []
        for j in range(1, qtdColuna+1):
            print('{:^5}'.format('[ ]'), end='')
            if cont in bombas:
                jogo.append('-1')
            elif cont not in bombas:
                jogo.append('0')
            cont += 1
        print()
        game.append(jogo)
    #print(game)

    gravarTabuleiro(calcularAdjacente(game))

def jogar():
    wb = load_workbook(filename='banco.xlsx', read_only=False)
    jogo = wb['jogo']
    maxLinha = jogo.max_row
    maxColuna = jogo.max_column

    gameOver = False
    ganhou = False
    jogadas = []
    
    dificuldade, qtdLinha, qtdColuna = configuracoes()
    
    while True:
        print('Insira as posições desejadas:')
        n1 = int(input('Linha: '))
        n2 = int(input('Coluna: '))
        # teste de jogadas
        
        if n1 > maxLinha or n2 > maxColuna:
            print('Jogada maior que o tabuleiro, tente novamente')
            print('Tamanho atual: {} linhas e {} colunas'.format(maxLinha, maxColuna))

        elif jogo.cell(column=n2, row=n1) not in jogadas:
            jogadas.append(jogo.cell(column=n2, row=n1))
            for i in range(1, maxLinha+1):
                for j in range(1, maxColuna + 1):
                    passouK = False
                    for k in jogadas:
                        # print(k, jogo.cell(column=j, row=i))
                        if k == jogo.cell(column=j, row=i) and jogo.cell(column=j, row=i).value != '-1':
                            print('{:^5}'.format(
                                jogo.cell(column=j, row=i).value), end='')
                            passouK = True
                            if len(jogadas) == (qtdColuna*qtdLinha - len(bombinhas())):
                                ganhou = True

                        elif k == jogo.cell(column=j, row=i) and jogo.cell(column=j, row=i).value == '-1':
                            print('{:^5}'.format('X'), end='')
                            passouK = True
                            gameOver = True
                    if passouK == False:
                        print('{:^5}'.format('[ ]'), end='')
                print('\n')
        else:
            print('Jogada já efetuada, tente novamente')

        if gameOver == True:
            print('Game Over!!!')
            wb.save('banco.xlsx')
            
        elif ganhou == True:
            print('Você Ganhou')
            wb.save('banco.xlsx')
            
  
        if ganhou == True or gameOver == True:
            escolha = input('Deseja Jogar Novamente?(S or N): ')
            if escolha.lower() == 's':
                print()
                menu()
            break



def bemVindo():
    print('''
         \|/        
        .-*-        
       / /|\        
      _L_              ____   U _____ u   __  __     __     __                   _   _       ____       U  ___ u
    ,°   °.         U | __°)u \| ___°|/ U|° \/ °|u   \ \   /°/u       ___       | \ |°|     |  _°\       \/°_ \/ 
(\ /  O O  \ /)      \|  _ \/  |  _|°   \| |\/| |/    \ \ / //       |_°_|     <|  \| |>   /| | | |      | | | | 
 \|    _    |/        | |_) |  | |___    | |  | |     /\ V /_,-.      | |      U| |\  |u   U| |_| |\ .-,_| |_| |
   \  (_)  /          |____/   |_____|   |_|  |_|    U  \_/-(_/     U/| |\ u    |_| \_|     |____/ u  \_)-\___/
   _/.___,\_         _|| \ \_   <<   >>  <<,-,,-.       //        .-,_|___|_,-.  ||   \ \,-.   |||_         \ \ 
  (_/ alf \_)       (__) (__) (__) (__)  (./  \.)     (__)        \_)-° °-(_/   (_°)  (_/   (__)_)         (__)
=====================================================================================================================
                               Bem vindo ao campo minado do Paulo e da Narumi =)
''')

def menu():
    dificuldade, linhas, colunas = configuracoes()

    print('''escolha uma das opções abaixo
    1 - JOGO
    2 - CONFIGURAÇÕES
    3 - SAIR''')
    n = input('Escolha: ')

    if n == '1': 
        criarTabuleiro(linhas,colunas)
        jogar()
    elif n == '2':
        alterarConfiguracao()
    elif n == '3':
        print('sair')
    else:
        print('Opção inválida\n')
        menu()

bemVindo()
menu()
